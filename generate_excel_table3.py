import pandas as pd
import numpy as np

# Data from Table 3 - Branch Comparison
data = []

# Direct LoTT (477 cases)
direct_lott_data = [
    ["Direct LoTT", "Gemini 2.0 Flash", "LoTT Framework", 0.644, 0.616, 0.720],
    ["Direct LoTT", "2.0 Flash", "Legal Syllogism", 0.577, 0.576, 0.601],
    ["Direct LoTT", "2.0 Flash", "CoT", 0.558, 0.560, 0.519],
    ["Direct LoTT", "2.0 Flash", "Standard", 0.520, 0.526, 0.399],
    ["Direct LoTT", "2.5 Flash", "Legal Syllogism", 0.637, 0.652, 0.709],
    ["Direct LoTT", "2.5 Flash", "Standard", 0.558, 0.557, 0.600],
    ["Direct LoTT", "2.5 Flash", "CoT", 0.551, 0.551, 0.607],
    ["Direct LoTT", "2.5 Pro", "Standard", 0.637, 0.659, 0.717],
    ["Direct LoTT", "2.5 Pro", "Legal Syllogism", 0.591, 0.648, 0.709],
    ["Direct LoTT", "2.5 Pro", "CoT", 0.614, 0.644, 0.710],
]

# RQ (23 cases)
rq_data = [
    ["RQ", "Gemini 2.0 Flash", "LoTT Framework", 0.696, 0.673, 0.588],
    ["RQ", "2.0 Flash", "CoT", 0.739, 0.670, 0.500],
    ["RQ", "2.0 Flash", "Legal Syllogism", 0.609, 0.591, 0.471],
    ["RQ", "2.0 Flash", "Standard", 0.652, 0.589, 0.429],
    ["RQ", "2.5 Flash", "CoT", 0.565, 0.642, 0.545],
    ["RQ", "2.5 Flash", "Legal Syllogism", 0.391, 0.444, 0.364],
    ["RQ", "2.5 Flash", "Standard", 0.348, 0.411, 0.348],
    ["RQ", "2.5 Pro", "Standard", 0.478, 0.589, 0.500],
    ["RQ", "2.5 Pro", "CoT", 0.478, 0.589, 0.500],
    ["RQ", "2.5 Pro", "Legal Syllogism", 0.391, 0.528, 0.462],
]

# Combine all data
all_data = direct_lott_data + rq_data

# Create DataFrame
df = pd.DataFrame(all_data, columns=[
    'Branch', 'Model', 'Method', 'Accuracy', 'Macro F1', 'Positive F1'
])

# Create Excel file with formatting
with pd.ExcelWriter('final_500_cases_evaluation_results_LAST_477_and_23/table3_branch_comparison.xlsx', 
                    engine='openpyxl') as writer:
    
    # Write main data
    df.to_excel(writer, sheet_name='Branch Comparison', index=False, startrow=2)
    
    # Get the workbook and worksheet
    workbook = writer.book
    worksheet = writer.sheets['Branch Comparison']
    
    # Add title and subtitle
    worksheet['A1'] = 'Table 3: Branch Comparison Analysis'
    worksheet['A2'] = 'Direct LoTT (477 cases) vs RQ (23 cases) - Performance Metrics'
    
    # Format headers
    from openpyxl.styles import Font, PatternFill, Alignment
    
    # Title formatting
    worksheet['A1'].font = Font(bold=True, size=14)
    worksheet['A2'].font = Font(italic=True, size=11)
    
    # Header formatting
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col in range(1, 7):  # A to F columns
        cell = worksheet.cell(row=3, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Format data cells
    for row in range(4, len(all_data) + 4):
        for col in range(1, 7):
            cell = worksheet.cell(row=row, column=col)
            cell.alignment = Alignment(horizontal="center")
            
            # Highlight LoTT Framework rows
            if worksheet.cell(row=row, column=3).value == "LoTT Framework":
                cell.fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
                cell.font = Font(bold=True)
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width

print("✓ Excel file created: table3_branch_comparison.xlsx")
print("✓ File location: final_500_cases_evaluation_results_LAST_477_and_23/")
print("✓ Contains formatted data with:")
print("  - Direct LoTT Branch (477 cases): 10 method combinations")
print("  - RQ Branch (23 cases): 10 method combinations") 
print("  - Metrics: Accuracy, Macro F1, Positive F1")
print("  - LoTT Framework rows highlighted in blue")